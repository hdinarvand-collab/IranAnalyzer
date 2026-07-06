"""
Excel Report V8 PRO FIXED - TradingView Style + Education Layer
Iran Analyzer
"""

import os
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.formatting.rule import ColorScaleRule


class ExcelReport:

    def __init__(self, symbol, technical, support, volume,
                 smart_money, divergence, pattern, final_score, df):

        self.symbol = symbol
        self.technical = technical
        self.support = support
        self.volume = volume
        self.smart_money = smart_money
        self.divergence = divergence
        self.pattern = pattern
        self.final_score = final_score
        self.df = df

        self.wb = Workbook()
        self.wb.remove(self.wb.active)

        self.output_dir = "Reports"
        os.makedirs(self.output_dir, exist_ok=True)

    # ======================================
    # STYLE
    # ======================================

    def header(self, cell):
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E79")
        cell.alignment = Alignment(horizontal="center")

    # ======================================
    # DASHBOARD
    # ======================================

    def dashboard(self):

        ws = self.wb.create_sheet("Dashboard")
        ws.freeze_panes = "A2"

        ws["A1"] = f"IRAN ANALYZER V8 PRO - {self.symbol}"
        ws["A1"].font = Font(size=16, bold=True, color="1F4E79")

        ws["A3"] = "Final Score"
        ws["B3"] = self.final_score.get("final_score", 0)
        ws["C3"] = "امتیاز نهایی"

        ws["A4"] = "Signal"
        ws["B4"] = self.final_score.get("signal", "")
        ws["C4"] = "سیگنال نهایی"

        ws["A5"] = "Date"
        ws["B5"] = datetime.now().strftime("%Y-%m-%d %H:%M")
        ws["C5"] = "زمان تحلیل"

        # KPIs
        kpis = [
            ("RSI", self.technical.get("rsi_signal", "")),
            ("EMA", self.technical.get("ema_signal", "")),
            ("MACD", self.technical.get("macd_status", "")),
            ("ADX", self.technical.get("adx_signal", "")),
            ("Trend", self.technical.get("trend", "")),
            ("Position", self.technical.get("price_position", "")),
        ]

        row = 7
        for k, v in kpis:
            ws[f"A{row}"] = k
            ws[f"B{row}"] = v
            ws[f"C{row}"] = "Indicator"
            row += 1

        # COLOR SIGNAL SAFE
        signal = self.final_score.get("signal", "")

        color_map = {
            "خرید بسیار قوی": "00C853",
            "خرید": "7CFC00",
            "نگهداری": "FFD700",
            "ضعیف": "FFA500",
            "فروش": "FF0000"
        }

        ws["B4"].fill = PatternFill(
            "solid",
            fgColor=color_map.get(signal, "FFFFFF")
        )

        return ws

    # ======================================
    # SIGNALS
    # ======================================

    def signals(self):

        ws = self.wb.create_sheet("Signals")
        ws.freeze_panes = "A2"

        ws.append(["Indicator", "Signal", "Score", "Description"])

        for c in ws[1]:
            self.header(c)

        data = [
            ("RSI", self.technical.get("rsi_signal", ""), 8,
             "قدرت خرید/فروش"),

            ("EMA", self.technical.get("ema_signal", ""), 10,
             "جهت روند"),

            ("MACD", self.technical.get("macd_status", ""), 10,
             "مومنتوم"),

            ("ADX", self.technical.get("adx_signal", ""), 8,
             "قدرت روند"),

            ("Ichimoku", self.technical.get("price_position", ""), 12,
             "ساختار روند"),
        ]

        for row in data:
            ws.append(row)

        return ws

    # ======================================
    # SCORE BREAKDOWN
    # ======================================

    def score(self):

        ws = self.wb.create_sheet("Score Breakdown")
        ws.freeze_panes = "A2"

        ws.append(["Component", "Score", "Meaning"])

        for c in ws[1]:
            self.header(c)

        breakdown = [
            ("RSI", 8, "قدرت خرید/فروش"),
            ("EMA", 10, "جهت روند"),
            ("MACD", 10, "مومنتوم"),
            ("ADX", 8, "قدرت روند"),
            ("Ichimoku", 12, "ساختار روند"),
            ("Support", self.support.get("support_score", 0), "سطوح حمایتی"),
            ("Volume", self.volume.get("volume_score", 0), "حجم معاملات"),
            ("Smart Money", self.smart_money.get("smart_money_score", 0), "پول هوشمند"),
            ("Divergence", self.divergence.get("divergence_score", 0), "واگرایی"),
            ("Pattern", self.pattern.get("pattern_score", 0), "الگو"),
        ]

        for row in breakdown:
            ws.append(row)

        # SAFE FIX (ColorScaleRule correct args)
        rule = ColorScaleRule(
            start_type="min", start_color="FF0000",
            mid_type="percentile", mid_value=50, mid_color="FFFF00",
            end_type="max", end_color="00C853"
        )

        ws.conditional_formatting.add("B2:B50", rule)

        return ws

    # ======================================
    # RAW DATA (KEEP)
    # ======================================

    def raw(self):

        ws = self.wb.create_sheet("Raw Data")
        ws.freeze_panes = "A2"

        ws.append(list(self.df.columns))

        for r in self.df.itertuples(index=False):
            ws.append(r)

        return ws

    # ======================================
    # CHARTS SAFE FIX
    # ======================================

    def charts(self, ws):

        max_row = min(len(self.df), len(self.df) + 1)

        # PRICE
        chart1 = LineChart()
        chart1.title = "Price Trend"

        data = Reference(ws, min_col=5, min_row=1, max_row=max_row)
        chart1.add_data(data, titles_from_data=True)

        ws.add_chart(chart1, "H2")

        # VOLUME
        chart2 = BarChart()
        chart2.title = "Volume"

        vdata = Reference(ws, min_col=6, min_row=1, max_row=max_row)
        chart2.add_data(vdata, titles_from_data=True)

        ws.add_chart(chart2, "H20")

    # ======================================
    # EXPORT FIXED
    # ======================================

    def export(self):

        print("🚀 Excel V8 PRO STARTED")

        try:

            self.dashboard()
            self.signals()
            self.score()

            raw = self.raw()
            self.charts(raw)

            filename = f"{self.symbol}_V8_PRO_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
            path = os.path.join(self.output_dir, filename)

            self.wb.save(path)

            print("=================================")
            print("✅ EXCEL CREATED SUCCESSFULLY")
            print("📁 FILE:")
            print(path)
            print("=================================")

        except Exception as e:
            print("❌ V8 ERROR:")
            print(e)